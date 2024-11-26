
from ast import Raise
from cgitb import text
from os import name
import re

import time
from tkinter import TkVersion, Variable, Widget, messagebox

from time import strftime
import sqlite3

import tkinter as tk
from tkinter import ttk
from turtle import color, width
#from networkx import max_weight_clique
from click import command
import openpyxl

# importing my first MODULE
from datetime import datetime

from sqlalchemy import ColumnClause, column, insert
#from ese_time import monday_e_time

import pygame

import pyttsx3


# initialize pygame sound 
pygame.mixer.init()
#initialize app
root = tk.Tk()
root.title("EseFlow.online")

root.configure(bg="black")
root.geometry("1150x550")

# here I set the custom color
def change_color():
    colornew = "#234575"
    return colornew

def change_color2():
    colornew2 = "#222"
    return colornew2

bg_colour = change_color()

#bg_colour = "#222"

#####************************************** 
#      This function will update the Treeview window 
#####***********************###************
def update_treeview_data():
    path = "expenses_tracker.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    treeview.delete(*treeview.get_children())
    
def clear_widgets(frame):
    for widget in frame.winfo_children():
        
        widget.destroy()
    
    
def load_frame5():
    # clear_widgets(frame1)
    # clear_widgets(frame2)
    #clear_widgets(frame4)
    frame5.tkraise()
    frame5.pack_propagate(False)

    # Here I added a Label to display text
    tk.Label(
        frame5,
        text=" INCOME TRACKER   ",
        bg=bg_colour,
         fg="Cyan",
        font=("TkMenuFont", 10)
        ).pack(padx=5, pady=5, )
    
    def create_table():
        conn = sqlite3.connect('ese_savings2.db')
        c = conn.cursor()
        c.execute('''CREATE TABLE IF NOT EXISTS ese_savings
                    (name TEXT, date TEXT, savings REAL, income REAL)''')
        conn.commit()
        conn.close()

    def add_contact():
        
        if name_enty.get():
            
            try:
                
                name = name_enty.get().capitalize()
                income = float(income_entry.get())
                fixedE = float(expenses_entry.get())
                variablE = float(variable_entry.get())
                incomeDiscription = income_dis.get()
                expensesDiscription = expenses_disc.get()
                variableDiscription = var_disc.get()
                
                date2 = datetime.now().strftime(f" %d-%m-%Y   %H:%M:%p   ")
                date = date2
                
                total_expenses = fixedE + variablE
                
                savings = income - total_expenses
                
                
                # writing to a text file 
                # Here I added all transaction into text file 
                #%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
                
                dba = open("dbaz.txt", "a")
                dba.write(f"-----    {date2}  --------- \n ")
                dba.write("\n")
                dba.write(f"*** HELLO {name.upper()} **** \n ")
                dba.write(f" *** On this day being *** {date2} \n ".upper())
                dba.write(f" ************************ \n")
                dba.write(f" **** HERE IS YOUR FINANCIAL STATMENT BASE YOUR ENTRIES: ***** \n")
                dba.write(f" ************************ \n")
                dba.write(f" Income:                = ${income}   from -------- {incomeDiscription}  \n")
                dba.write(f" Fixed Expenditure:     = ${fixedE}   for  ---------  {expensesDiscription} \n")
                dba.write(f" Variable Expenditure:  = ${variablE} for ------------{variableDiscription}  \n")
                dba.write(f" Total Expenditure:     = ${total_expenses} \n")
                dba.write(f" *** ------------------***\n")
                dba.write(f"        Balance *********************[[]] = ${savings} [[]] \n")
                dba.write("\n")
                dba.write("\n")
                dba.close()
                
                        #%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%    
                        
                
                
                # writing to a text file 
                # Here I added all transaction into text file 
                #%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
                dba = open(f"{name}.txt", "w")
                dba.write(f"-----    {date2}  --------- \n ")
                dba.write("\n")
                dba.write(f"*** HELLO {name.upper()} **** \n ")
                dba.write(f" *** On this day being *** {date2} \n ".upper())
                dba.write(f" ************************ \n")
                dba.write(f" **** HERE IS YOUR FINANCIAL STATMENT BASE YOUR ENTRIES: ***** \n")
                dba.write(f" ************************ \n")
                dba.write(f" Income:                = ${income}   from --------  {incomeDiscription} \n")
                dba.write(f" Fixed Expenditure:     = ${fixedE}   for  --------- {expensesDiscription}  \n")
                dba.write(f" Variable Expenditure:  = ${variablE}   for  ----------{variableDiscription}   \n")
                dba.write(f" Total Expenditure:     = ${total_expenses} \n")
                dba.write(f" *** ------------------***\n")
                dba.write(f"        Balance ************************[[]] = ${savings} [[]] \n")
                dba.write("\n")
                dba.write("\n")
                dba.close()
                
                
                # writing to a text file 
                # Here I added all transaction into text file 
                #%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%
                dba = open(f"dbaz_two.txt", "w")
                dba.write(f"-----    {date2}  --------- \n ")
                dba.write("\n")
                dba.write(f"*** HELLO {name.upper()} **** \n ")
                dba.write(f" *** On this day being *** {date2} \n ".upper())
                dba.write(f" ************************ \n")
                dba.write(f" **** HERE IS YOUR FINANCIAL STATMENT BASE YOUR ENTRIES: ***** \n")
                dba.write(f" ************************ \n")
                dba.write(f" Income:                = ${income}   from --------  {incomeDiscription} \n")
                dba.write(f" Fixed Expenditure:     = ${fixedE}   for  --------- {expensesDiscription}  \n")
                dba.write(f" Variable Expenditure:  = ${variablE}   for  ----------{variableDiscription}   \n")
                dba.write(f" Total Expenditure:     = ${total_expenses} \n")
                dba.write(f" *** ------------------***\n")
                dba.write(f"        Balance ***********************[[]] = ${savings} [[]] \n")
                dba.write("\n")
                dba.write("\n")
                dba.close()
                
                
                global show_current_savings
                def show_current_savings():
                    
                    
                    global current_balance
                    current_balance = tk.Listbox(frame5, bg="#222", fg="white", width=8, height=1, border=5, font=("TkHeadFont", 9))
                    current_balance.place(x=160, y=40)    
                    #current_balance.insert(0, "$")
                    current_balance.insert(0, savings)
            
                
                show_current_savings()
                
                        #%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%    
            # savings = phone_ent.get()
                conn = sqlite3.connect('ese_savings2.db')
                c = conn.cursor()
                c.execute("INSERT INTO ese_savings2 (name, date, savings, income) VALUES (?, ?, ?, ?)", (name, date, savings, income))
                conn.commit()
                conn.close()
                
#(((((((((((((((((((((((((((((((((())))))))))))))))))))))))))))))))))
            #  Treewiew code starts here 
#**********************************************************************
              
                path = "expenses_tracker.xlsx"     
                workbook = openpyxl.load_workbook(path)
                sheet = workbook.active
                row_value = [name, income, fixedE, variablE, savings, date2]
                sheet.append(row_value)
                workbook.save(path)

                update_treeview_data()
                
                load_treeview_data()
                
                show_total_savings()
                
                signal_red()
                
                all_balance()

#(((((((((((((((((((((((((((((((((((())))))))))))))))))))))))))))))))))))
         
                
                name_enty.delete(0, "end")
                name_enty.insert(0, "Name")
                
                income_entry.delete(0, "end")
                income_entry.insert(0, "Income")
                
                expenses_entry.delete(0, "end")
                expenses_entry.insert(0, "Fixed Exp")
                
                variable_entry.delete(0, "end")
                variable_entry.insert(0, " Variable Expenses") 
                
                income_dis.delete(0, "end")
                income_dis.insert(0, " I Discription" )
                
                expenses_disc.delete(0, "end")
                expenses_disc.insert(0, " E Discription")
                
                var_disc.delete(0, "end")
                var_disc.insert(0, "Var Discription") 

            except Exception as e:
                messagebox.showerror("Error", str(e))
                
                name_enty.delete(0, "end")
                name_enty.insert(0, "Name")
                
                income_entry.delete(0, "end")
                income_entry.insert(0, "Income")
                
                expenses_entry.delete(0, "end")
                expenses_entry.insert(0, "Fixed Exp")
                
                variable_entry.delete(0, "end")
                variable_entry.insert(0, " Variable Expenses") 
                
                income_dis.delete(0, "end")
                income_dis.insert(0, " I Discription" )
                
                expenses_disc.delete(0, "end")
                expenses_disc.insert(0, " E Discription")
                
                var_disc.delete(0, "end")
                var_disc.insert(0, "Var Discription") 
                
                
                # ***********************************************
                
 
        else:
            messagebox.showerror("Error", "Please fill in the entry box to add data")  
            
            name_enty.delete(0, "end")
            name_enty.insert(0, "Name")
            
            income_entry.delete(0, "end")
            income_entry.insert(0, "Income")
            
            expenses_entry.delete(0, "end")
            expenses_entry.insert(0, "Fixed Exp")
            
            variable_entry.delete(0, "end")
            variable_entry.insert(0, " Variable Expenses") 
            
            income_dis.delete(0, "end")
            income_dis.insert(0, " I Discription" )
            
            expenses_disc.delete(0, "end")
            expenses_disc.insert(0, " E Discription")
            
            var_disc.delete(0, "end")
            var_disc.insert(0, "Var Discription") 
        
    def show_contacts():
        conn = sqlite3.connect('ese_savings2.db')
        c = conn.cursor()
        c.execute("SELECT * FROM ese_savings2")
        rows = c.fetchall()
        conn.close()
        list_view2.delete(0, tk.END)
        for row in rows:
            list_view2.insert(tk.END, row)
            last_income.insert(tk.END, row)
              
            
           # print(row[1])
            
      
         
    def show_total_savings():
        conn = sqlite3.connect('ese_savings2.db')
        c = conn.cursor()
        
         # total monthly_sale 
        Total_saved= "select total(savings) from ese_savings2"
        c.execute(Total_saved) 
        global total_display
        total_display = (c.fetchone()[0])
        conn.close()
        
    show_total_savings()
    
    
    def all_balance():
        global last_income
        last_income = tk.Listbox(frame5, bg="#222", fg="white", width=8, height=1, border=5, font=("TkHeadFont", 9))
        last_income.place(x=290, y=40)    
        #last_income.insert(0, "$")
        last_income.insert(0, total_display)
        
        # Here I added sound
        
        # pygame.mixer.music.load("audio/good.mp3")
        # pygame.mixer.music.play(loops=0)
        
            #^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^    
        global total_savings_sh
        total_savings_sh = tk.Listbox(frame5, bg="#222", fg="white", width=10, height=1, border=5, font=("TkHeadFont", 9))
        total_savings_sh.place(x=400, y=40)    
        #total_savings_sh.insert(0, "$")
        total_savings_sh.delete(0, "end")
        total_savings_sh.insert(0, total_display)
        
        
        if total_display >=1:
            
                # Here I added sound
            pygame.mixer.music.load("audio/successful.wav")
            pygame.mixer.music.play(loops=0)
            
        else:
                # Here I added sound
            pygame.mixer.music.load("audio/warning.wav")
            pygame.mixer.music.play(loops=0)
            
    
    all_balance()
    
    
    #show_current_savings()
  
    

#  This code will display a red button when the savings is less than 1
#  to indicate that there is sortage of money in the account 
    
    def signal_red():
        if total_display>=1:
            red_button =tk.Button(
                frame5,
                text="S",
                bg="green",
                fg="white",
                font=("TkHeadFont", 8),
                cursor= "hand2",
                width=3,
                activebackground="#badee2",
                activeforeground="black",).place(x=423, y=10)
            
        else:
            red_button =tk.Button(
                frame5,
                text="R",
                bg="red",
                fg="white",
                font=("TkHeadFont", 8),
                cursor= "hand2",
                width=3,
                activebackground="#badee2",
                activeforeground="black",).place(x=423, y=10)
    
    signal_red()
    
    tk.Label(
        frame5,
        text="Bal.",
        bg=bg_colour,
        fg="cyan",
        font=("Microsoft Yahei UI Light", 11)
        ).place(x=370, y=40)
    #^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
    
    #***********************************************

    tk.Label(
        frame5,
        text="Inc.",
        bg=bg_colour,
        fg="cyan",
        font=("Microsoft Yahei UI Light", 11)
        ).place(x=260, y=40)
    #************************************************
        
         # clear list view  in frame 5    
    def delete_contacts():
        conn = sqlite3.connect('ese_savings2.db')
        c = conn.cursor()
        c.execute("DELETE FROM ese_savings2 WHERE name='Monday Eseinone'")
        rows = c.fetchall()
        conn.close()
        list_view2.delete(0, tk.END)
        for row in rows:
            list_view2.insert(tk.END, row)
          
        total_savings_sh.delete(0, tk.END)
        total_savings_sh.insert(0, total_display)
            

    #@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@    
    # def select():
    #     print()
    #     print( " hellow world ")
    #@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
           
    tk.Label(
        frame5,
        text=" ",
        bg=bg_colour,
       # bg="black",
        fg="white",
        width=2,
        height=0,
        font=("TkMenuFont", 2)
        ).pack(padx=5, pady=30)
    
    tk.Label(
        frame5,
        text="NAME",
       # bg=bg_colour,
        bg=bg_colour,
        width=13,
        fg="white",
        font=("Microsoft Yahei UI Light", 10)
        ).place(x=15, y=90)

    tk.Label(
        frame5,
        text=" DATE  ",
        bg=bg_colour,
        fg="white",
        width=13,
        font=("Microsoft Yahei UI Light", 10)
        ).place(x=130, y=90)
    
    tk.Label(
        frame5,
        text=" TIME  ",
        bg=bg_colour,
        fg="white",
        width=13,
        font=("Microsoft Yahei UI Light", 10)
        ).place(x=230, y=90)

    tk.Label(
        frame5,
        text=" SAVINGS ",
        bg=bg_colour,
        fg="white",
        width=13,
        font=("Microsoft Yahei UI Light", 10)
        ).place(x=340, y=90)

    
#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
    def on_enterv(e):
        name_enty.delete(0, "end")

    def on_leavev(e):
        name = name_enty.get()
        if name=="":
            name_enty.insert(0,"  Username") 
            
    name_enty = tk.Entry(frame5, bg="#675645", fg="white", width=12,border=4, font=("TkHeadFont", 12))
    name_enty.place(x=10, y=15)
    name_enty.insert(0, "  Username")
    name_enty.bind("<FocusIn>", on_enterv)
    name_enty.bind("<FocusOut>", on_leavev)

#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@  
    # Income Entry starts here 
#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
    def on_enterdd(e):
        income_entry.delete(0, "end")

    def on_leavedd(e):
        incomeN = income_entry.get()
        if incomeN=="":
            income_entry.insert(0," Income") 
            
    income_entry = tk.Entry(frame5, bg="#675645", fg="white", width=10,border=4, font=("TkHeadFont", 12))
    income_entry.place(x=85, y=380)
    income_entry.insert(0, " Income")
    income_entry.bind("<FocusIn>", on_enterdd)
    income_entry.bind("<FocusOut>", on_leavedd)
#-------------------------------------------------
    
    tk.Label(
        frame5,
        text="INCOME ",
        bg=bg_colour,
        fg="white",
     
        font=("ds-digital",12)
        ).place(x=10, y=385)

# ***************DISCRIPTION ENTRY FOR INCOME  ******************
    def on_enter(e):
        income_dis.delete(0, "end")

    def on_leave(e):
        incomeK = income_dis.get()
        if incomeK=="":
            income_dis.insert(0," Discription") 
            
    income_dis = tk.Entry(frame5, bg="#675645", fg="white", width=12,border=4, font=("TkHeadFont", 12))
    income_dis.place(x=230, y=380)
    income_dis.insert(0, " Discription")
    income_dis.bind("<FocusIn>", on_enter)
    income_dis.bind("<FocusOut>", on_leave)
#-------------------------------------------------
    
    tk.Label(
        frame5,
        text="D ",
        bg=bg_colour,
        fg="white",
     
        font=("ds-digital",12)
        ).place(x=205, y=385)
#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@  
   # Expneses Entry starts here 
#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
    def aon_enterb(e):
        expenses_entry.delete(0, "end")

    def aon_leaveb(e):
        exp_n = expenses_entry.get()
        if exp_n=="":
            expenses_entry.insert(0," Expenses") 
            
    expenses_entry = tk.Entry(frame5, bg="#675645", fg="white", width=10,border=4, font=("TkHeadFont", 12))
    expenses_entry.place(x=85, y=420)
    expenses_entry.insert(0, "  Expenses ")
    expenses_entry.bind("<FocusIn>", aon_enterb)
    expenses_entry.bind("<FocusOut>", aon_leaveb)
#-------------------------------------------------
    
    tk.Label(
        frame5,
        text="EXP ",
        bg=bg_colour,
        fg="white",
     
        font=("ds-digital",12)
        ).place(x=10, y=420)
    
#*************** DISCRIPTION ENTRY FOR EXPENSES  
#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
    def on_enterb(e):
        expenses_disc.delete(0, "end")

    def on_leaveb(e):
        exp_d1 = expenses_disc.get()
        if exp_d1=="":
            expenses_disc.insert(0," Expenses") 
            
    expenses_disc = tk.Entry(frame5, bg="#675645", fg="white", width=12,border=4, font=("TkHeadFont", 12))
    expenses_disc.place(x=230, y=420)
    expenses_disc.insert(0, "  Expenses ")
    expenses_disc.bind("<FocusIn>", on_enterb)
    expenses_disc.bind("<FocusOut>", on_leaveb)
#-------------------------------------------------
    
    tk.Label(
        frame5,
        text="D ",
        bg=bg_colour,
        fg="white",
     
        font=("ds-digital",12)
        ).place(x=205, y=420)
#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@ 
   # variable Expenses Entry starts here 
#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
    def on_enterbv(e):
        variable_entry.delete(0, "end")

    def on_leavebv(e):
        var_exp = variable_entry.get()
        if var_exp=="":
            variable_entry.insert(0," Variable Expenses") 
            
    variable_entry = tk.Entry(frame5, bg="#675645", fg="white", width=10,border=4, font=("TkHeadFont", 12))
    variable_entry.place(x=85, y=460)
    variable_entry.insert(0, "  Variable Expenses ")
    variable_entry.bind("<FocusIn>", on_enterbv)
    variable_entry.bind("<FocusOut>", on_leavebv)
#-------------------------------------------------
    
    tk.Label(
        frame5,
        text="EXP",
        bg=bg_colour,
        fg="white",
     
        font=("ds-digital",12)
        ).place(x=10, y=465)
    
    
#***************** DISCRIPTION FOR  VARIABLE EXPENSES *************
    def on_enterbvaa(e):
        var_disc.delete(0, "end")

    def on_leavebvaa(e):
        vpvar = var_disc.get()
        if vpvar=="":
            var_disc.insert(0," Variable Expenses") 
            
    var_disc = tk.Entry(frame5, bg="#675645", fg="white", width=12,border=4, font=("TkHeadFont", 12))
    var_disc.place(x=230, y=460)
    var_disc.insert(0, "  Variable Expenses ")
    var_disc.bind("<FocusIn>", on_enterbvaa)
    var_disc.bind("<FocusOut>", on_leavebvaa)
#-------------------------------------------------
    
    tk.Label(
        frame5,
        text="D",
        bg=bg_colour,
        fg="white",
     
        font=("ds-digital",12)
        ).place(x=205, y=465)
#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@
    
    def key_return(event):
            action = add_contact()
            tab_order()
            return action
            
    root.bind('<Return>', key_return)

    submite_button = tk.Button(
        frame5,
        text="SUBMIT ",
        bg="black",
        fg="white",
        font=("TkHeadFont", 10),
        cursor= "hand2",
        width=14,
        activebackground="#badee2",
        activeforeground="black", 
        command=lambda:key_return).place(x=230, y=500) 
    
    # binding the ENTER KEY TO FUNCTION. I got the idea from the website link bellow: 
    # https://discuss.python.org/t/keyboard-enter-key-to-function-as-a-submit-button/25354
    # def key_h(event):
    #     label["text"] = "You pressed h"

    #root.bind('<Return>', key_return)
   # app.bind('<h>', key_h)
    
   # '<Return>', on_entry
   
    # Change tab to order 
    name_enty.focus()
    
    def tab_order():
        
        if name_enty.focus():
            name_enty.focus()
            try:               
                Widgets = [name_enty, income_entry, income_dis, expenses_entry, expenses_disc, variable_entry, var_disc, submite_button ]
                
                for w in Widgets:
                    w.lift()
                    
            except Exception as e:
                messagebox.showerror("Error", str(e))
        #else:
          #  messagebox.showerror(f"{name}", "Click Enter key to Continue ") 

    list_scroll = tk.Scrollbar(frame5)
    list_scroll.pack(side="right", fill="y")
    
   
    list_view2 = tk.Listbox(frame5, bg="black", fg="red", font=("ds-digital", 12), width= 50, height=8, border=5)#  bg="#658765",
    list_view2.pack(padx=10, pady=8)
    
    list_scroll.config(command=list_view2.yview)


    tk.Button(
        frame5,
        text="Show Contact",
        bg="black",
        fg="white",
        font=("TkHeadFont", 10),
        cursor= "hand2",
        width=12,
        activebackground="#badee2",
        activeforeground="black",
        command=lambda:show_contacts()).place(x=370, y=380)
    
    tk.Button(
        frame5,
        text="Update Data",
        bg="black",
        fg="white",
        font=("TkHeadFont", 10),
        cursor= "hand2",
        width=12,
        activebackground="#badee2",
        activeforeground="black",
        command=lambda:delete_contacts()).place(x=370, y=420)
        # command=lambda:delete_contacts()).pack(pady= 5)
    
    create_table() 

    
    tk.Button(
        frame5,
        text="Total Savings ",
        bg="black",
        fg="white",
        font=("TkHeadFont", 10),
        cursor= "hand2",
        width=12,
        activebackground="#badee2",
        activeforeground="black",
        command=lambda:show_total_savings()).place(x=370, y=460)
   
    tk.Button(
        frame5,
        text="Expand WINDOW",
        bg="black",
        fg="white",
        font=("TkHeadFont", 8),
        cursor= "hand2",
        width=16,
        activebackground="#badee2",
        activeforeground="black",
        command=lambda:show_total_savings()).place(x=370, y=500)  
            
 
#&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&
#$$$$$$$$$$$$$$$$  ROOT LEBELS AND DISPLAY $$$$$$$$$$$$$$$$$$$$$$$$$$$$
#********************************************
#  This lable covers the top of the Treeview 
#********************************************
    tk.Label(root,
        text="",
        bg=bg_colour,
        fg="white",
        width=48,
        height=20,
        font=("ds-digital",22)
        ).place(x=505, y=0)

    
    tk.Label(root,
        text="EseFlow.online   Excel View  ",
        bg=bg_colour,
        fg="white",
        width=44,
        font=("ds-digital",14)
        ).place(x=530, y=5)
    
    
    color_button = tk.Button(
        root,
        text="COLOR ",
        bg="black",
        fg="white",
        font=("TkVersion", 12),
        cursor= "hand2",
        width=16,
        activebackground="#badee2",
        activeforeground="black",
    ).place(x=990, y=470)   # command=lambda:show_total_savings()
        
    file_button = tk.Button(
        root,
        text="OPEN",
        bg="black",
        fg="white",
        font=("TkVersion", 12),
        cursor= "hand2",
        width=16,
        activebackground="#badee2",
        activeforeground="black",
        command=lambda:open_text()).place(x=990, y=510) 
    
  #***********************************************
# col = { ( "Monday", "Eseinoe", "Income", "Expenses", "Variable_Expenses")}
# for i in col:
#     a = i
#     print(i)
#     col_b = a[:5]   
# last_income = tk.Listbox(root, bg="#222", fg="white", width=50, height=15, border=5, font=("TkHeadFont", 14))
# last_income.place(x=520, y=50)    
# #last_income.insert(0, "$")
# #last_income.insert(0, total_display)
# # last_income.insert(0, "welcome ")
# last_income.insert(tk.END, col_b)

#(((((((((((((((((((((((((((((((())))))))))))))))))))))))))))))))  
#  Treeview list starts here 
#*****************************************************************
    def load_treeview_data():
        path ="expenses_tracker.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        
        list_values = list(sheet.values)
        print(list_values)
        
        for col_name in list_values[0]:
            treeview.heading(col_name, text=col_name)
            
        
        for value_tuple in list_values[1:]:
            treeview.insert('', tk.END, values=value_tuple)
            
        
        
    tree_scroll = ttk.Scrollbar(root)
    tree_scroll.place(x=1100, y=50, height=200) # (side="right", fill="y")

    global treeview
    global cols_list
    cols_list = ("Name", "Income", "Fixed_Exp", "Var_Exp","Savings","Date")
    treeview = ttk.Treeview(root, show="headings",
                            yscrollcommand=tree_scroll.set, columns=cols_list, height=9)  
    treeview.column("Name", width=100)
    treeview.column("Income", width=60)
    treeview.column("Fixed_Exp", width=60)
    treeview.column("Var_Exp", width=70)
    treeview.column("Savings", width=60)
    treeview.column("Date", width=80)
    
    treeview.place(x=530, y=50, width=550)  
    
    tree_scroll.config(command=treeview.yview)
    
    load_treeview_data()

#(((((((((((((((((((((((((((((((((())))))))))))))))))))))))))))))))))  

   
    # Time function and Label starts here 
    def time():
        string = strftime("EseFlow.online Time: " + '%H:%M:%S %p     %d-%m-%y')
        label.config(text=string)
        label.after(1000, time)
        
    label = tk.Label(root, font=("ds-digital", 13), background= "black", foreground= "cyan" )
    label.place(x=740, y=510, anchor='center')
    time()
# (((((((((((((((((((((((((((((())))))))))))))))))))))))))))))
#      Author: by Monday Eseinone Completed 18th March, 2024. 
#####***********************###************ 
    autor = tk.Label(root, font=("ds-digital", 8), text="Created By Monday Eseinone", border=2, background= bg_colour, foreground= "white" )
    autor.place(x=530, y=520)
    
    author_email_address = tk.Label(root, font=("ds-digital", 8), text="Email: eeseinone@gmail.com", background= bg_colour, foreground= "yellow" )
    author_email_address.place(x=690, y=520)
    
    author_address = tk.Label(root, font=("ds-digital", 8), text="Contact: +234 8075236542", background= bg_colour, foreground= "red" )
    author_address.place(x=840, y=520)


#####************************************** 
#   My Text file start here 
#((((((((((((((((((((((((((((()))))))))))))))))))))))))))))

    def open_text():
        text_file = open("dbaz.txt", "r")
        stuff = text_file.read()
        
        my_textfile.insert("end", stuff)
        
        text_file.close()

    
    
    
    text_scroll= tk.Scrollbar(root)
    text_scroll.place(x=1100, y=270, height=190)
    
    my_textfile = tk.Text(root, width=60, height=10, bg="black", fg="cyan", font=("ds-digital", 12), border=5)
    my_textfile.place(x=530, y=270)
    text_scroll.config(command=my_textfile.yview)

            
frame5 = tk.Frame(root, width=500, height=550, bg=bg_colour)
frame5.grid(row=0, column=1)

frame_down = tk.Frame(root, width=500, height=100, bg=bg_colour)
frame_down.place(x=0, y=555)
frame_down.config(pady=0)


load_frame5()



root.mainloop()
